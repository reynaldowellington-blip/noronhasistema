import os  
import json  
import time  
import pandas as pd  
from PIL import Image  
from dotenv import load_dotenv  
from google import genai  
from google.genai import types  
from openpyxl import Workbook  
from openpyxl.styles import Font, PatternFill, Alignment  

load_dotenv()  

# Inicialização do Gemini  
chave_api = os.getenv("GEMINI_API_KEY")  
client = genai.Client(api_key=chave_api)  

PASTA_FOTOS = "."  
ARQUIVO_EXCEL_SAIDA = "cotacoes_odoo.xlsx"  
ARQUIVO_CLIENTES_ODOO = "clientesnoronha.xlsx"  

print("📂 Carregando base de clientes do Odoo...")
if os.path.exists(ARQUIVO_CLIENTES_ODOO):
    df_odoo_clientes = pd.read_excel(ARQUIVO_CLIENTES_ODOO)
    clientes_odoo_cadastrados = set(df_odoo_clientes['Cliente'].dropna().astype(str).str.strip().str.upper())
    print(f"✅ {len(clientes_odoo_cadastrados)} clientes carregados.")
else:
    clientes_odoo_cadastrados = set()
    print("⚠️ Arquivo clientesnoronha.xlsx não encontrado.")

vendas = []  
compras = []  

# Mapeia apenas as fotos de talões, ignorando arquivos de logo automaticamente
fotos = [f for f in os.listdir(PASTA_FOTOS) if f.lower().endswith(('.jpg', '.jpeg', '.png')) and 'logo' not in f.lower()]  
print(f"📸 Encontradas {len(fotos)} fotos de talões para processar.\n")  

prompt_instrucao = """  
Você é um assistente especializado em ler talões de pedidos manuscritos e notas fiscais.  
Analise a imagem e identifique se a operação é uma VENDA (bloco/talão emitido pela empresa) ou COMPRA (recibo/nota de fornecedor externo).  

Instruções de Extração:  
1. "tipo_operacao": OBRIGATORIAMENTE "VENDA" ou "COMPRA".  
2. "cliente_fornecedor": Nome do Cliente/Empresa ou CPF/CNPJ. Se não houver, coloque "Cliente Não Identificado".  
3. "data_compra": Data no formato DD/MM/AAAA.  
4. "forma_pagamento": PIX, Dinheiro, Cartão, A Vista, etc.  
5. "itens": Lista contendo "produto", "quantidade" e "preco_unitario".  

Retorne APENAS o JSON puro no seguinte formato:  
{  
 "tipo_operacao": "VENDA",  
 "cliente_fornecedor": "Nome ou CNPJ/CPF",  
 "data_compra": "28/08/2026",  
 "forma_pagamento": "PIX",  
 "itens": [  
 {  
 "produto": "Nome do Produto",  
 "quantidade": 1.0,  
 "preco_unitario": 0.00  
 }  
 ]  
}  
"""  

for foto in fotos:  
    foto_path = os.path.join(PASTA_FOTOS, foto)  
    print(f"🔍 Lendo foto com IA: {foto}...")  
       
    for tentativa in range(3):  
        try:  
            imagem = Image.open(foto_path)  
            response = client.models.generate_content(  
                model='gemini-2.0-flash',  
                contents=[imagem, prompt_instrucao],  
                config=types.GenerateContentConfig(response_mime_type="application/json")  
            )  
               
            texto_resposta = response.text.strip()  
            if texto_resposta.startswith("```json"):  
                texto_resposta = texto_resposta.replace("```json", "", 1)  
            if texto_resposta.startswith("```"):  
                texto_resposta = texto_resposta.replace("```", "", 1)  
            if texto_resposta.endswith("```"):  
                texto_resposta = texto_resposta[:-3].strip()  
               
            dados = json.loads(texto_resposta)  
            tipo = dados.get("tipo_operacao", "VENDA").upper()  
            parceiro_original = str(dados.get("cliente_fornecedor", "Não Identificado")).strip()  
            
            parceiro_upper = parceiro_original.upper()
            if parceiro_upper in clientes_odoo_cadastrados:
                parceiro_final = parceiro_original  
            else:
                encontrou = False
                for c_odoo in clientes_odoo_cadastrados:
                    if parceiro_upper in c_odoo or c_odoo in parceiro_upper:
                        parceiro_final = c_odoo
                        encontrou = True
                        break
                if not encontrou:
                    parceiro_final = f"{parceiro_original} (cliente novo)"

            data_doc = dados.get("data_compra", "")  
            forma_pgto = dados.get("forma_pagamento", "")  
               
            for item in dados.get("itens", []):  
                qtd = float(item.get("quantidade", 1.0))  
                preco_unit = float(item.get("preco_unitario", 0.00))  
                registro = {  
                    "Cliente/Fornecedor": parceiro_final,  
                    "Data": data_doc,  
                    "Forma Pagamento": forma_pgto,  
                    "Produto": item.get("produto", ""),  
                    "Quantidade": qtd,  
                    "Preço Unitário": preco_unit,  
                    "Total": qtd * preco_unit,  
                }  
                if tipo == "VENDA":  
                    vendas.append(registro)  
                else:  
                    compras.append(registro)  
            break  
        except Exception as e:  
            print(f"Tentativa {tentativa+1} falhou para {foto}: {e}")  
            time.sleep(2)  

with pd.ExcelWriter(ARQUIVO_EXCEL_SAIDA, engine='openpyxl') as writer:
    df_vendas = pd.DataFrame(vendas) if vendas else pd.DataFrame(columns=["Cliente/Fornecedor", "Data", "Forma Pagamento", "Produto", "Quantidade", "Preço Unitário", "Total"])
    df_compras = pd.DataFrame(compras) if compras else pd.DataFrame(columns=["Cliente/Fornecedor", "Data", "Forma Pagamento", "Produto", "Quantidade", "Preço Unitário", "Total"])
    
    df_vendas.to_excel(writer, sheet_name='Vendas', index=False)
    df_compras.to_excel(writer, sheet_name='Compras', index=False)

from openpyxl import load_workbook
wb = load_workbook(ARQUIVO_EXCEL_SAIDA)

ws_vendas = wb['Vendas']
fill_vendas_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")

for cell in ws_vendas[1]:
    cell.fill = fill_vendas_header
    cell.font = font_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_compras = wb['Compras']
fill_compras_header = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")

for cell in ws_compras[1]:
    cell.fill = fill_compras_header
    cell.font = font_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(ARQUIVO_EXCEL_SAIDA)
print(f"\n🚀 Processo concluído! Planilha gerada com sucesso: {ARQUIVO_EXCEL_SAIDA}")